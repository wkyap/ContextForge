"""Reports API — dashboard metrics, at-risk detection, SSG compliance reports."""

from __future__ import annotations

import io
from datetime import UTC, datetime
from typing import Any

from fastapi import APIRouter, Query
from fastapi.responses import JSONResponse, StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from contextforge.api.deps import PostgresDep

router = APIRouter(prefix="/reports")


@router.get("/dashboard")
async def dashboard_metrics(
    postgres: PostgresDep,
    programme_id: str | None = None,
) -> dict[str, Any]:
    """Real-time programme KPI metrics for coordinator dashboard."""
    where = ""
    params: list[Any] = []
    if programme_id:
        where = "WHERE programme_id::text = $1"
        params = [programme_id]

    # Trainee counts by status
    status_rows = await postgres.fetch(
        f"SELECT status, count(*) as cnt FROM trainees {where} GROUP BY status",
        params,
    )
    status_counts = {r["status"]: r["cnt"] for r in status_rows}

    # Total trainees
    total = sum(status_counts.values())
    placed = status_counts.get("placed", 0)
    completed = status_counts.get("completed", 0) + placed
    placement_rate = (placed / completed * 100) if completed > 0 else 0

    # Pending verifications
    pending_docs = await postgres.fetch_one(
        "SELECT count(*) as cnt FROM documents WHERE verification_status = 'pending'",
        [],
    )

    # Pending enrolment reviews
    pending_apps = await postgres.fetch_one(
        "SELECT count(*) as cnt FROM applications WHERE status = 'pending'",
        [],
    )

    return {
        "kpi": {
            "total_trainees": total,
            "placement_rate": round(placement_rate, 1),
            "pending_verifications": pending_docs["cnt"] if pending_docs else 0,
            "pending_enrolments": pending_apps["cnt"] if pending_apps else 0,
        },
        "trainee_status": status_counts,
        "active_trainees": status_counts.get("training", 0),
        "completed": completed,
        "placed": placed,
    }


@router.get("/at-risk")
async def at_risk_trainees(postgres: PostgresDep) -> dict[str, Any]:
    """Identify trainees at risk of non-completion or non-placement."""
    # Trainees completed >3 months ago without placement
    rows = await postgres.fetch(
        """SELECT t.id, t.trainee_code, t.name, t.programme_type, t.status,
           t.updated_at, t.email
           FROM trainees t
           WHERE t.status = 'completed'
           AND t.updated_at < now() - interval '90 days'
           AND NOT EXISTS (
               SELECT 1 FROM placements p
               WHERE p.trainee_id = t.id AND p.status = 'verified'
           )
           ORDER BY t.updated_at ASC LIMIT 50""",
        [],
    )

    at_risk = []
    for r in rows:
        at_risk.append({
            **dict(r),
            "risk_reason": "Completed training >90 days ago without verified placement",
            "recommended_action": "Send placement reminder; review matching opportunities",
        })

    return {"at_risk_count": len(at_risk), "trainees": at_risk}


@router.get("/placement-funnel")
async def placement_funnel(postgres: PostgresDep) -> dict[str, Any]:
    """Placement pipeline funnel data."""
    status_rows = await postgres.fetch(
        "SELECT status, count(*) as cnt FROM trainees GROUP BY status ORDER BY status",
        [],
    )

    funnel_order = ["applied", "enrolled", "training", "completed", "placed"]
    funnel = {}
    for status in funnel_order:
        funnel[status] = 0
    for r in status_rows:
        if r["status"] in funnel:
            funnel[r["status"]] = r["cnt"]

    return {"funnel": funnel}


async def _ssg_data(postgres: PostgresDep) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    programmes = await postgres.fetch(
        "SELECT * FROM programmes WHERE active = true", []
    )
    metrics = await postgres.fetch(
        """SELECT p.name as programme_name, p.type as programme_type,
           count(DISTINCT t.id) as total_trainees,
           count(DISTINCT CASE WHEN t.status = 'completed' THEN t.id END) as completed,
           count(DISTINCT CASE WHEN t.status = 'placed' THEN t.id END) as placed,
           count(DISTINCT pl.id) as verified_placements
           FROM programmes p
           LEFT JOIN trainees t ON t.programme_id = p.id
           LEFT JOIN placements pl ON pl.trainee_id = t.id AND pl.status = 'verified'
           GROUP BY p.id, p.name, p.type""",
        [],
    )
    return [dict(p) for p in programmes], [dict(m) for m in metrics]


def _write_ssg_workbook(programmes: list[dict[str, Any]], metrics: list[dict[str, Any]]) -> bytes:
    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F46E5")
    center = Alignment(horizontal="center", vertical="center")

    def _autosize(ws: Any) -> None:
        for col_cells in ws.columns:
            length = max((len(str(c.value)) if c.value is not None else 0)
                         for c in col_cells)
            ws.column_dimensions[get_column_letter(col_cells[0].column)].width = (
                min(max(length + 2, 12), 40)
            )

    # Sheet 1 — Cover
    cover = wb.active
    assert cover is not None
    cover.title = "Cover"
    cover["A1"] = "SSG Monthly Compliance Report"
    cover["A1"].font = Font(bold=True, size=16)
    cover["A3"] = "Generated"
    cover["B3"] = datetime.now(UTC).isoformat(timespec="seconds")
    cover["A4"] = "Active programmes"
    cover["B4"] = len(programmes)
    cover["A5"] = "Total trainees"
    cover["B5"] = sum(int(m.get("total_trainees") or 0) for m in metrics)
    cover["A6"] = "Verified placements"
    cover["B6"] = sum(int(m.get("verified_placements") or 0) for m in metrics)
    cover["A8"] = "Sign-off"
    cover["A8"].font = Font(bold=True)
    cover["A9"] = "Approver"
    cover["A10"] = "Date"
    cover["A11"] = "Signature"
    _autosize(cover)

    # Sheet 2 — Programmes
    ws = wb.create_sheet("Programmes")
    if programmes:
        cols = list(programmes[0].keys())
        for j, c in enumerate(cols, 1):
            cell = ws.cell(row=1, column=j, value=c)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
        for i, p in enumerate(programmes, 2):
            for j, c in enumerate(cols, 1):
                v = p.get(c)
                ws.cell(row=i, column=j,
                        value=v.isoformat() if hasattr(v, "isoformat") else v)  # type: ignore[union-attr]
        _autosize(ws)

    # Sheet 3 — Metrics
    ws = wb.create_sheet("Metrics")
    cols = ["programme_name", "programme_type", "total_trainees",
            "completed", "placed", "verified_placements", "placement_rate_%"]
    for j, c in enumerate(cols, 1):
        cell = ws.cell(row=1, column=j, value=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
    for i, m in enumerate(metrics, 2):
        completed = int(m.get("completed") or 0)
        placed = int(m.get("placed") or 0)
        rate = round((placed / completed * 100), 1) if completed else 0.0
        row = [m.get("programme_name"), m.get("programme_type"),
               int(m.get("total_trainees") or 0), completed, placed,
               int(m.get("verified_placements") or 0), rate]
        for j, v in enumerate(row, 1):
            ws.cell(row=i, column=j, value=v)
    _autosize(ws)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@router.post("/ssg", response_model=None)
async def generate_ssg_report(
    postgres: PostgresDep,
    programme_id: str | None = None,
    fmt: str = Query("json", pattern="^(json|xlsx)$"),
) -> JSONResponse | StreamingResponse:
    """Generate SSG compliance report — JSON payload or Excel workbook."""
    programmes, metrics = await _ssg_data(postgres)

    if fmt == "xlsx":
        data = _write_ssg_workbook(programmes, metrics)
        ts = datetime.now(UTC).strftime("%Y%m%d-%H%M%S")
        return StreamingResponse(
            io.BytesIO(data),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="ssg-report-{ts}.xlsx"',
            },
        )

    return JSONResponse(
        content={
            "report_type": "ssg_monthly",
            "requires_signoff": True,
            "programmes": programmes,
            "metrics": metrics,
            "download_xlsx": "POST /api/v1/reports/ssg?fmt=xlsx",
        }
    )
